import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')
SRC = r"D:\박홍석\광진소방서 IP전자전화기(전화번호) 일람표(2026.1.12.).xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f"=== {sh} ({ws.max_row}x{ws.max_column}) ===")
    for r in range(1, min(ws.max_row + 1, 8)):
        row = []
        for c in range(1, min(ws.max_column + 1, 15)):
            v = ws.cell(r, c).value
            row.append(f"{c}:{repr(v)[:20]}")
        print(" r", r, " | ".join(row))
    print()
