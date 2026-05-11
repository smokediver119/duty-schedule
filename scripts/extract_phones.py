"""Extract 성명 <-> 경비(내선번호) mapping from 일람표 sheet."""
import json
import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

SRC = r"D:\박홍석\광진소방서 IP전자전화기(전화번호) 일람표(2026.1.12.).xlsx"
OUT = Path("seed/phones.json")

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["일람표"]

# Left section: name col 5, ext col 6
# Right section: name col 13, ext col 14
sections = [(5, 6), (13, 14)]

records = []
seen = set()
for r in range(4, ws.max_row + 1):
    for name_col, ext_col in sections:
        name = ws.cell(r, name_col).value
        ext = ws.cell(r, ext_col).value
        if not name or ext is None:
            continue
        try:
            ext_s = str(int(ext))
        except (TypeError, ValueError):
            ext_s = str(ext).strip()
        name_s = str(name).strip()
        if not name_s or not ext_s.isdigit():
            continue
        key = (name_s, ext_s)
        if key in seen:
            continue
        seen.add(key)
        records.append({"name": name_s, "ext": ext_s})

OUT.parent.mkdir(exist_ok=True)
OUT.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"✅ {len(records)} records → {OUT}")

# spot check
check = ["김용열", "남궁명", "최원길", "박종범", "박홍석"]
for c in check:
    hit = [r for r in records if r["name"] == c]
    print(f"  {c}: {hit}")
