"""Extract 근무자 데이터 and 2026년 5월 당직표 from xlsx into JSON seeds."""
import json
import openpyxl
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

XLSX = "2026년 5월 (상황)당직근무 지정(초안).xlsx"
OUT_DIR = Path("seed")
OUT_DIR.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)

# --- Users ---
ws = wb["근무자 데이터"]
users = []
for row in ws.iter_rows(min_row=2, values_only=True):
    dept, rank, name = row[1], row[2], row[3]
    if not name:
        continue
    # Role mapping per planning doc
    rank_s = str(rank).strip()
    if rank_s == "소방령":
        role = "special"  # 특별경계 (rule.md)
    elif rank_s == "소방경":
        role = "supervisor"  # 당직 책임관
    elif rank_s in ("소방위", "소방장"):
        role = "leader"  # 조장
    else:  # 소방교, 소방사
        role = "member"  # 조원
    users.append({
        "name": str(name).strip(),
        "dept": str(dept).strip(),
        "rank": rank_s,
        "role": role,
        "active": True,
    })

# assign stable orderIndex by role in input order
for role in ("supervisor", "leader", "member", "special"):
    idx = 0
    for u in users:
        if u["role"] == role:
            u["orderIndex"] = idx
            idx += 1

(OUT_DIR / "users.json").write_text(
    json.dumps(users, ensure_ascii=False, indent=2), encoding="utf-8"
)

# --- Duty schedule (2026년 5월) ---
ws = wb["당직상황근무 지정 현황"]
duties = {}
current_day = None

for r in range(8, ws.max_row + 1):
    b = ws.cell(r, 2).value  # 일자
    c = ws.cell(r, 3).value  # 구분
    d = ws.cell(r, 4).value  # 요일
    sup_name = ws.cell(r, 6).value
    lead_name = ws.cell(r, 8).value
    mem_name = ws.cell(r, 10).value

    if b:
        m = re.match(r"5월\s*(\d+)일", str(b))
        if m:
            current_day = int(m.group(1))

    if not current_day or not sup_name:
        continue

    iso = f"2026-05-{current_day:02d}"
    shift = str(c).strip() if c else ""  # 일직 / 숙직 / ""
    weekday = str(d).strip() if d else None

    if iso not in duties:
        duties[iso] = {
            "date": iso,
            "weekday": weekday,
            "type": "weekday",
            "assignments": [],
        }

    # weekend/holiday has 일직/숙직 split
    if shift in ("일직", "숙직"):
        duties[iso]["type"] = "weekend_or_holiday"
        shift_en = "day" if shift == "일직" else "night"
    else:
        shift_en = "full"  # 평일 24시간

    duties[iso]["assignments"].append({
        "shift": shift_en,
        "supervisor": str(sup_name).strip(),
        "leader": str(lead_name).strip() if lead_name else None,
        "member": str(mem_name).strip() if mem_name else None,
    })

(OUT_DIR / "duties_2026_05.json").write_text(
    json.dumps(list(duties.values()), ensure_ascii=False, indent=2), encoding="utf-8"
)

print(f"✅ users: {len(users)} → seed/users.json")
print(f"✅ duties: {len(duties)} days → seed/duties_2026_05.json")
print()
print("Role counts:")
from collections import Counter
c = Counter(u["role"] for u in users)
for k, v in c.items():
    print(f"  {k}: {v}")
